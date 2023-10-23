
import os
import re
import yaml
import random
import string
import datetime
import openpyxl
import platform

def get_current_directory():
    """
    获取当前文件目录
    """
    current_directory = os.getcwd()
    return current_directory

def mkdir(directory_name):
    """
    支持递归创建文件夹
    """
    new_directory_path = os.path.join("", directory_name)
    if not os.path.exists(new_directory_path):
        os.makedirs(new_directory_path)
        #print(f"Created directory: {new_directory_path}")
    else:
        pass
        #print(f"Directory already exists: {new_directory_path}")

def create_directory(directory_name):
    """
    支持递归创建文件夹
    """
    new_directory_path = os.path.join("", directory_name)
    if not os.path.exists(new_directory_path):
        os.makedirs(new_directory_path)
        #print(f"Created directory: {new_directory_path}")
    else:
        pass
        #print(f"Directory already exists: {new_directory_path}")

def list_directory_contents(directory_name):
    """
    查看文件夹内所有文件
    """
    modified_files = []

    for root, dirs, files in os.walk(directory_name):
        for file in files:
            file_path = os.path.join(root, file)
            modified_files.append(file_path)
    return modified_files

def delete_file(file_name):
    """
    删除文件
    """
    file_path = os.path.join("", file_name)
    if os.path.exists(file_path) and os.path.isfile(file_path):
        os.remove(file_path)
        print(f"Deleted file: {file_path}")
    else:
        print(f"File not found: {file_path}")

def search_keyword_files(directory, keyword):
    """
    搜索文件夹内所有文件跟关键字有关的文件
    xxx dir keyword
    parser = argparse.ArgumentParser(description='在指定目录中搜索关键字')
    parser.add_argument('directory', type=str, help='待检测目录的路径')
    parser.add_argument('keyword', type=str, help='要搜索的关键字')
    args = parser.parse_args()
    directory, keyword = args.directory, args.keyword
    """
    for root, _, files in os.walk(directory):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            try:
                with open(file_path, 'rb') as file:
                    line_number = 0
                    for line in file:
                        line_number += 1
                        if keyword in line.decode('utf-8', 'ignore'):
                            #print(f"File: {file_path}, Line: {line_number}, Content: {line.strip().decode('utf-8', 'ignore')}")
                            print(f"File: {file_path}, Line: {line_number}")
            except Exception as e:
                print(f"Error reading {file_path}: {str(e)}")

def clear():
    """Clear the screen"""
    os_name = platform.system()
    if os_name == "Windows":
        os.system('cls')
    else:
        os.system('clear')

def read_config_yaml_file():
    """
    读取配置文件的逻辑
    """
    try:
        with open("config.yaml", 'r', encoding='gb18030', errors='ignore') as file:
            data = yaml.safe_load(file)
        return data
    except FileNotFoundError:
        # 处理文件不存在的情况
        print(f'找不到配置文件 config.yaml\n')
        return None
    except yaml.YAMLError:
        # 处理 YAML 解析错误的情况
        print(f'配置文件 config.yaml 解析错误\n')
        return None

def read_host_file(filename) -> list:
    """
    读取 host txt 文件的逻辑
    """
    li = []
    try:
        with open(filename, 'r', encoding='gb18030', errors='ignore') as f:
            lines = f.readlines()
    except Exception as E:
        # 处理 YAML 解析错误的情况
        print(f'{filename}读取错误, 或非txt文件 {E}\n')
        return li

    # windows上可能出现这类情况，所以将非标准的换行符替换成标准的""
    lines = [line.replace("\r\n", "") for line in lines]

    # Linux
    lines = [line.replace("\n", "") for line in lines]

    for line in lines:
        li.append(line)
    return li

def read_poc_yaml_file(file_path):
    """
    读取 poc yaml 文件的逻辑
    """
    try:
        with open(file_path, 'r', encoding='UTF-8', errors='ignore') as file:
            data = yaml.safe_load(file)
            print(f"{file_path}......加载成功")
    except FileNotFoundError:
        # 处理文件不存在的情况
        print(f'POC {file_path} 不存在\n')
        return None
    except yaml.YAMLError:
        # 处理 YAML 解析错误的情况
        print(f'POC {file_path} 解析错误\n')
        return None

    if data:
        return data


def ip_terr(text):
    """
    提取ip、port
    """
    # 定义匹配IP地址和可选端口号的正则表达式模式
    ip_with_optional_port_pattern = r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(?::\d+)?'

    # 使用正则表达式找到匹配的IP地址和可选端口号
    ip_ports = re.findall(ip_with_optional_port_pattern, text)

    # 打印提取到的IP地址和端口号
    for ip_port in ip_ports:
        url = ip_port.spilt(':')
        if len(url) < 2:
            host, port = url[0], 80
        else:
            host, port = url[0], url[1]
        return (host, port)

def write_to_xlsx(ti: list, data: list, filename: str, flag=1):
    """
    写入xlsx文件
    """
    if flag == 1:
        workbook = openpyxl.Workbook()
        del workbook["Sheet"]
        ws1 = workbook.create_sheet("汇总", 1)
    else:
        workbook = openpyxl.load_workbook(filename)
        ws1 = workbook.create_sheet("脆弱", 2)
    ws1.append(ti)
    for d in data:
        try:
            ws1.append(d)
        except Exception as E:
            pass
    workbook.save(filename)
    if flag == 1:
        print(f'导出文件 {filename}\n')

def getime():
    """
    获取当前时间
    """
    current_datetime = datetime.datetime.now()
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H-%M-%S")
    return formatted_datetime


def rand_word() -> str:
    """
    获取随机6位字符串
    """
    letters = string.ascii_letters  # 包含所有字母的字符串（大小写）
    random_letters = ''.join(random.choice(letters) for _ in range(6))

    return str(random_letters)
